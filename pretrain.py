import os
import re
import torch
import torch.nn as nn
from torch.nn import functional as F
import numpy as np
import logging
import argparse
import time
import json
import datetime
from pathlib import Path
from apex import amp
from apex.parallel import DistributedDataParallel as DDP
from deap import gp
from models import select_config, select_single_model, nas_bert_models
from models.PKD import PKD
from tokenizers import select_basic_tokenizer
from utils import AverageMeter, register_custom_ops, register_custom_ops3, set_seeds, setup_logger, get_entire_linear_idx, \
    get_entire_params, calc_params, reduce_tensor, save_checkpoint, load_pretrain_state_dict, load_resume_state_dict, \
    load_multi_task_state_dict, load_supernet_state_dict_6_540_to_12_360, load_supernet_state_dict_6_540_to_6_360, \
    create_optimizer, create_scheduler, create_pretrain_dataset
from models.bert import MySupernetBertAttention
from models.mobile_bert import MobileBertAttention

parser = argparse.ArgumentParser()
parser.add_argument('--distributed', action='store_true', help='distributed mode')
parser.add_argument('--fp16', action='store_true', help='mixed precision training mode')
parser.add_argument('--opt_level', default='O1', type=str, help='fp16 optimization level')
parser.add_argument('--gpu_devices', default='4,5,6,7', type=str, help='available gpu devices')
parser.add_argument('--seed', default=42, type=int, help='seed')

parser.add_argument('--lowercase', action='store_true', help='whether to do lowercase')
parser.add_argument('--temperature', default=1, type=float, help='temperature for soft cross entropy loss')
parser.add_argument('--ffn_expr', default='[]', type=str, help='feed-forward network expression')
parser.add_argument('--dynamic_heads', default=0, type=int, help='whether cut heads by PCA')
parser.add_argument('--teacher_model', default='bert_base', type=str, help='teacher model name')
parser.add_argument('--student_model', default='tiny_bert', type=str, help='student model name')
parser.add_argument('--vocab_path', default='', type=str, help='path to pretrained vocabulary file')
parser.add_argument('--merge_path', default='', type=str, help='path to pretrained merge file (for roberta)')
parser.add_argument('--train_ratio', default=1, type=float, help='ratio of train dataset')
parser.add_argument('--val_ratio', default=0, type=float, help='ratio of val dataset')

parser.add_argument('--start_epoch', default=1, type=int, help='start epoch (default is 1)')
parser.add_argument('--total_epochs', default=3, type=int, help='total epochs')
parser.add_argument('--batch_size', default=64, type=int, help='batch size')
parser.add_argument('--lr', default=1e-4, type=float, help='initial learning rate')
parser.add_argument('--optim_type', default='adamw', type=str, help='optimizer type')
parser.add_argument('--sched_type', default='step', type=str, help='lr scheduler type')
parser.add_argument('--warmup_proportion', default=0.1, type=float, help='proportion of warmup steps')
parser.add_argument('--momentum', default=0.9, type=float, help='SGD momentum')
parser.add_argument('--weight_decay', default=0.01, type=float, help='weight decay')
parser.add_argument('--max_grad_norm', default=1, type=float, help='max gradient norm')
parser.add_argument('--disp_freq', default=50, type=int, help='display frequency')
parser.add_argument('--save_freq', default=1, type=int, help='checkpoint save frequency')
parser.add_argument('--ckpt_keep_num', default=1, type=int, help='max number of checkpoint files to keep')

parser.add_argument('--teacher_pretrain_path', default='', type=str,help='path to pretrained state dict of teacher model')
parser.add_argument('--student_pretrain_path', default='', type=str, help='path to pretrained student state dict')
parser.add_argument('--student_multi_task_pretrain_path', default='', type=str, help='path to multi-task pretrained student state dict')
parser.add_argument('--student_pretrain_path_6_540_to_6_360', default='', type=str, help='path to pretrained student state dict from L6-H540 to L6-H360')
parser.add_argument('--student_pretrain_path_6_540_to_12_360', default='', type=str, help='path to pretrained student state dict from L6-H540 to L12-H360')
parser.add_argument('--student_resume_path', default='', type=str, help='path to resume checkpoint of student model')
parser.add_argument('--wiki_dir', default='', type=Path, help='directory to wikipedia dataset')
parser.add_argument('--book_dir', default='', type=Path, help='directory to bookcorpus dataset')
parser.add_argument('--concate_data_dir', default='', type=Path, help='directory to concatenated dataset')
parser.add_argument('--exp_dir', default='./exp/tmp/', type=str, help='experiment directory')
parser.add_argument('--local_rank', default=0, type=int, help='DDP local rank')
parser.add_argument('--world_size', default=1, type=int, help='DDP world size')
parser.add_argument('--type_blocks', default=3, type=int, help='number of the type')
parser.add_argument('--test', default=0, type=int, help='is test')
parser.add_argument('--num_layers', default=12, type=int, help='')
parser.add_argument('--average_add_times', default=3, type=int, help='')
parser.add_argument('--dynamic', default=0, type=int, help='dynamic linear')
parser.add_argument('--range1', default=0, type=int, help='')
parser.add_argument('--range2', default=0, type=int, help='')
parser.add_argument('--type_each_block', default=1, type=int, help='')
parser.add_argument('--new_loss', default=1, type=int, help='')
parser.add_argument('--sample_steps', default=1000, type=int, help='')
parser.add_argument('--upper_params', default=15700000, type=float, help='max params to search')
parser.add_argument('--calc_freq', default=30, type=int, help='pca/cka/sim calc frequency')
parser.add_argument('--min_pca', default=1, type=int, help='pca/cka/sim calc frequency')
parser.add_argument('--cka_upper_limit', default=0.9, type=float, help='SGD momentum')
parser.add_argument('--only2', action='store_true', help='if only 2 blocks work')
parser.add_argument('--get_avg_cka', action='store_true', help='')

args = parser.parse_args()

os.environ['CUDA_VISIBLE_DEVICES'] = args.gpu_devices
print(args.gpu_devices)

from pathlib import PosixPath
if args.book_dir!=PosixPath('.') and args.concate_data_dir!=PosixPath('.'):
    args.all_train_dir = [args.wiki_dir, args.book_dir]
else:
    args.all_train_dir = [args.wiki_dir]

print(args)
print(args.wiki_dir, args.book_dir)
print(args.all_train_dir)

def main():
    args.exp_dir = os.path.join(args.exp_dir, datetime.datetime.now().strftime('%Y%m%d-%H%M%S'))
    setup_logger(args.exp_dir)
    if args.local_rank == 0:
        logging.info(args)

    use_gpu = False
    if args.gpu_devices and torch.cuda.is_available():
        use_gpu = True
    if use_gpu and args.local_rank == 0:
        logging.info('Currently using GPU: {}'.format(args.gpu_devices))
    elif not use_gpu and args.local_rank == 0:
        logging.info('Currently using CPU')
    set_seeds(args.seed, use_gpu)

    if use_gpu and args.distributed:
        torch.cuda.set_device(args.local_rank)
        torch.distributed.init_process_group(backend='nccl', init_method='env://')
        args.world_size = torch.distributed.get_world_size()
        logging.info('Training in distributed mode (process {}/{})'.format(args.local_rank + 1, args.world_size))

    # Register custom operators for auto bert
    if args.student_model in nas_bert_models:
        if args.student_model == 'auto_tiny_bert':
            pset = register_custom_ops3()
        else:
            pset = register_custom_ops()
        if args.student_model == 'auto_bert_12':
            args.ffn_expr = np.reshape([[x, x] for x in args.ffn_expr], -1).tolist()
        entire_ffn_func = [gp.compile(expr, pset) for expr in args.ffn_expr]
        entire_linear_idx = get_entire_linear_idx(args.ffn_expr)
        args.ffn_arch = [entire_ffn_func, entire_linear_idx]

    # Load model and tokenizer
    expr = eval(args.ffn_expr)
    #logging.info("expr={}".format(expr))
    if expr == []:
        issingle = args.range1 == args.range2
        args.type_blocks = (args.range2 - args.range1 + 1) * args.type_each_block
        args.range1 *= args.type_each_block
        args.range2 = (args.range2 + 1) * args.type_each_block - 1
    else:
        issingle = True
        args.range2 = 0
        args.range1 = 10000
        for e in expr:
            if e != expr[0]:
                issingle = False
            if args.range2 < e:
                args.range2 = e
            if args.range1 > e:
                args.range1 = e
        args.type_blocks = args.range2 - args.range1 + 1
        if args.student_model in ['mysupernet','mysupernet_10M','mysupernet_5M']:
            args.type_blocks = 12
    logging.info("range1={}, range2={}".format(args.range1, args.range2))
    teacher_config = select_config(args.teacher_model, args.lowercase, issingle, args.dynamic == 0)
    student_config = select_config(args.student_model, args.lowercase, issingle, args.dynamic == 0)
    teacher_model = select_single_model(args.teacher_model, args.lowercase, issingle=issingle, fixed_dimension=args.dynamic == 0)
    student_model = select_single_model(args.student_model, args.lowercase, issingle=issingle, fixed_dimension=args.dynamic == 0)
    args.tokenizer = select_basic_tokenizer(args.teacher_model, args.lowercase, args.vocab_path, args.merge_path)
    args.teacher_interval = teacher_config.num_layers // student_config.num_layers
    args.num_student_layers = student_config.num_layers
    args.upper_params = student_config.upper_params
    args.student_config = student_config

    if use_gpu:
        teacher_model, student_model = teacher_model.cuda(), student_model.cuda()
    if args.local_rank == 0:
        logging.info('Teacher model size: {:.2f}M'.format(calc_params(teacher_model) / 1e6))
        logging.info('Student model size: {:.2f}M'.format(calc_params(student_model) / 1e6))
        if args.student_model in nas_bert_models:
            logging.info('Student sub model size: {:.2f}M'.format(get_entire_params(student_config.param_list, args.ffn_expr)))
        logging.info('Student model config: {}'.format(args.student_config.__dict__))

    # Count total training examples
    all_num_data_epochs = []
    total_examples = 0
    for train_dir in args.all_train_dir:
        num_epoch_examples = []
        num_data_epochs = len([file for file in os.listdir(train_dir) if re.match(r'epoch_\d+_metrics.json', file) is not None])
        all_num_data_epochs.append(num_data_epochs)
        for i in range(num_data_epochs):
            metrics_file = train_dir / 'epoch_{}_metrics.json'.format(i)
            if metrics_file.is_file():
                metrics = json.loads(metrics_file.read_text())
                num_epoch_examples.append(metrics['num_training_examples'])
        for epoch in range(args.total_epochs):
            total_examples += int(num_epoch_examples[epoch % len(num_epoch_examples)] * args.train_ratio)

    for data_epoch in all_num_data_epochs:
        assert data_epoch == all_num_data_epochs[0]
    args.num_data_epochs = all_num_data_epochs[0]

    # Create optimizer and scheduler
    num_sched_steps = total_examples // (args.batch_size * args.world_size)
    num_warmup_steps = int(num_sched_steps * args.warmup_proportion)
    optimizer = create_optimizer(student_model, args.optim_type, args.lr, args.weight_decay, args.momentum)
    scheduler = create_scheduler(optimizer, args.sched_type, num_sched_steps, num_warmup_steps)

    # Enable fp16/distributed training
    if use_gpu:
        if args.fp16:
            amp.register_half_function(torch, 'einsum')
            student_model, optimizer = amp.initialize(student_model, optimizer, opt_level=args.opt_level)
            if args.local_rank == 0:
                logging.info('Using fp16 training mode')
        if args.distributed:
            teacher_model = DDP(teacher_model, delay_allreduce=True)
            student_model = DDP(student_model, delay_allreduce=True)
        else:
            teacher_model = nn.DataParallel(teacher_model)
            student_model = nn.DataParallel(student_model)

    # Load model weights
    ckpt_path = args.teacher_pretrain_path
    if ckpt_path:
        if os.path.exists(ckpt_path):
            load_pretrain_state_dict(args.teacher_model, teacher_model, ckpt_path, use_gpu)
            if args.local_rank == 0:
                logging.info('Loaded teacher pretrained state dict from \'{}\''.format(ckpt_path))
        else:
            if args.local_rank == 0:
                logging.info('Teacher pretrained state dict is not found in \'{}\''.format(ckpt_path))

    ckpt_path = args.student_pretrain_path
    if ckpt_path:
        if os.path.exists(ckpt_path):
            load_pretrain_state_dict(args.student_model, student_model, ckpt_path, use_gpu, is_finetune=True)
            if args.local_rank == 0:
                logging.info('Loaded student pretrained state dict from \'{}\''.format(ckpt_path))
        else:
            if args.local_rank == 0:
                logging.info('Student pretrained state dict is not found in \'{}\''.format(ckpt_path))

    ckpt_path = args.student_multi_task_pretrain_path
    if ckpt_path:
        if os.path.exists(ckpt_path):
            task_id = 0
            load_multi_task_state_dict(student_model, ckpt_path, task_id, is_finetune=True, load_pred=False)
            if args.local_rank == 0:
                logging.info('Loaded student multi-task pretrained state dict from \'{}\''.format(ckpt_path))
        else:
            if args.local_rank == 0:
                logging.info('Student multi-task pretrained state dict is not found in \'{}\''.format(ckpt_path))

    ckpt_path = args.student_pretrain_path_6_540_to_6_360
    if ckpt_path:
        if os.path.exists(ckpt_path):
            load_supernet_state_dict_6_540_to_6_360(student_model, ckpt_path)
            if args.local_rank == 0:
                logging.info('Loaded student pretrained state dict 6-540 to 6-360 from \'{}\''.format(ckpt_path))
        else:
            if args.local_rank == 0:
                logging.info('Student pretrained state dict 6-540 to 6-360 is not found in \'{}\''.format(ckpt_path))

    ckpt_path = args.student_pretrain_path_6_540_to_12_360
    if ckpt_path:
        if os.path.exists(ckpt_path):
            load_supernet_state_dict_6_540_to_12_360(student_model, ckpt_path)
            if args.local_rank == 0:
                logging.info('Loaded student pretrained state dict 6-540 to 12-360 from \'{}\''.format(ckpt_path))
        else:
            if args.local_rank == 0:
                logging.info('Student pretrained state dict 6-540 to 12-360 is not found in \'{}\''.format(ckpt_path))

    ckpt_path = args.student_resume_path
    if ckpt_path:
        if os.path.exists(ckpt_path):
            checkpoint = load_resume_state_dict(student_model, ckpt_path, optimizer, scheduler)
            args.start_epoch = checkpoint['epoch'] + 1
            if args.local_rank == 0:
                logging.info('Loaded student resume checkpoint from \'{}\''.format(ckpt_path))
                logging.info('Start epoch: {}'.format(args.start_epoch))
        else:
            if args.local_rank == 0:
                logging.info('Student resume checkpoint is not found in \'{}\''.format(ckpt_path))
    final_heads = []
    if args.dynamic_heads > 0:
        assert expr != []
        layer_sim, cka_pairs = train(teacher_model, student_model, optimizer, scheduler, use_gpu, get_cka_pair = True)
        if args.local_rank == 0:
            for i in range(len(cka_pairs)):
                cka_pairs[i] = cka_pairs[i].cpu().numpy()
            if args.get_avg_cka:
                non_one_count = []
                mean_value = []
                for i in range(len(cka_pairs)):
                    non_one_count.append(np.count_nonzero(cka_pairs[i]!=1))
                    mean_value.append(np.sum(cka_pairs[i][cka_pairs[i]!=1]) / non_one_count[i])
                logging.info("non_one_count={}".format(non_one_count))
                logging.info("mean_value={}".format(mean_value))
                logging.info("layer_sim={}".format(layer_sim))
                return
            #logging.info('pair_output={}'.format(pair_output))
            for i in range(args.num_student_layers):
                remove_heads = []
                for j in range(student_config.num_attn_heads):
                    if j in remove_heads:
                        continue
                    for k in range(j+1, student_config.num_attn_heads):
                        if cka_pairs[i][j][k] > args.cka_upper_limit:
                            remove_heads.append(k)
                final_heads.append(student_config.num_attn_heads-len(remove_heads))
                hidden_size = student_config.hidden_size
                if expr[i] >= args.type_each_block:
                    hidden_size = int(0.25 * hidden_size)
                #logging.info('expr[i]={}'.format(expr[i]))
                while hidden_size % final_heads[i] != 0:
                    final_heads[i] -= 1
                #if final_heads[i] == 6:
                #    logging.info("cka_pairs[i]={}".format(cka_pairs[i]))
                #    return
                student_model.module.encoder[i][expr[i]].attention.num_attn_heads = final_heads[i]
                student_model.module.encoder[i][expr[i]].attention.attn_head_size = hidden_size // final_heads[i]
                student_model.module.encoder[i][expr[i]].attention.all_head_size = student_model.module.encoder[i][expr[i]].attention.attn_head_size * final_heads[i]
            logging.info('final_heads={}'.format(final_heads))
        #student_model.modules().
    try:
        train(teacher_model, student_model, optimizer, scheduler, use_gpu)
    except KeyboardInterrupt:
        print('Keyboard interrupt (process {}/{})'.format(args.local_rank + 1, args.world_size))


def train(teacher_model, student_model, optimizer, scheduler, use_gpu, get_cka_pair = False):
    st_time = time.time()
    if args.local_rank == 0:
        logging.info('==> Start training')

    import openpyxl

    # 设置参数
    n = args.type_blocks
    m = args.num_student_layers

    # 初始化工作簿和工作表
    wb = openpyxl.Workbook()
    pca_ws = [wb.active] * n
    cka_ws = [wb.active] * n
    sim_ws = [wb.active] * n

    # 新建n张pca工作表
    for i in range(n):
        pca_ws[i] = wb.create_sheet('pca_' + str(i))

    # 新建n张cka工作表
    for i in range(n):
        cka_ws[i] = wb.create_sheet('cka_' + str(i))

    # 新建n张sim工作表
    for i in range(n):
        sim_ws[i] = wb.create_sheet('sim_' + str(i))

    for epoch in range(args.start_epoch, args.total_epochs + 1):
        train_dataset, train_loader = create_pretrain_dataset(
            args.all_train_dir, epoch - 1, args.tokenizer, args.num_data_epochs, args.local_rank, args.batch_size,
            use_gpu, args.distributed, 'train', args.train_ratio, args.val_ratio, args.concate_data_dir)

        if args.distributed:
            train_loader.sampler.set_epoch(epoch)

        if args.test != 1:
            if get_cka_pair == False:
                cka_each_layer, pca_each_layer, sim_each_layer = train_epoch(teacher_model, student_model, epoch, optimizer, scheduler, train_loader, use_gpu)
                if args.local_rank == 0:
                    logging.info('epoch = {}, cka_each_layer = {}, pca_each_layer = {}, sim_each_layer = {}'.format(epoch, cka_each_layer, pca_each_layer, sim_each_layer))
            else:
                layer_sim, cka_pairs = train_epoch(teacher_model, student_model, epoch, optimizer, scheduler, train_loader, use_gpu, get_cka_pair=get_cka_pair)
                return layer_sim, cka_pairs

            # 写入数据
            for i in range(n):
                for j in range(m):
                    pca_ws[i].cell(column=j + 1, row=epoch).value = pca_each_layer[i][j]
                    cka_ws[i].cell(column=j + 1, row=epoch).value = cka_each_layer[i][j]
                    sim_ws[i].cell(column=j + 1, row=epoch).value = sim_each_layer[i][j]

            # 保存工作簿
            wb.save(os.path.join(args.exp_dir, "result.xlsx"))

        if epoch % args.save_freq == 0 or epoch == args.total_epochs:
            if args.local_rank == 0:
                state = {'state_dict': student_model.state_dict(),
                         'optimizer': optimizer.state_dict(),
                         'scheduler': scheduler.state_dict(),
                         'epoch': epoch}
                ckpt_name = 'ckpt_ep' + str(epoch) + '.bin'
                save_checkpoint(state, args.exp_dir, ckpt_name, args.ckpt_keep_num)
                logging.info('Student state dict has been saved to \'{}\''.format(os.path.join(args.exp_dir, ckpt_name)))
        if args.local_rank == 0:
            logging.info('-' * 50)
        if args.test == 1:
            #pass
            break

    if args.local_rank == 0:
        elapsed = round(time.time() - st_time)
        elapsed = str(datetime.timedelta(seconds=elapsed))
        logging.info('Finished, total training time (h:m:s): {}'.format(elapsed))

def attention_condfidence_normalized(outputs):
    metric_array = []
    for output in outputs:
        metric_array.append(torch.mean(torch.max(output, 1)[0]))

    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    summed = torch.tensor(0.0).to(device)
    for j in range(len(metric_array)):
        summed += torch.nansum(metric_array[j])
    summed /= len(metric_array)

    return summed.detach().item()

def attention_condfidence(outputs):
    metric_array = []
    for output in outputs:
        metric_array.append(torch.mean(torch.max(output, 1)[0]))

    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    summed = torch.tensor(0.0).to(device)
    for j in range(len(outputs)):
        summed += torch.nansum(metric_array[j])

    return summed.detach().item()

def train_epoch(teacher_model, student_model, epoch, optimizer, scheduler, data_loader, use_gpu, get_cka_pair = False):
    teacher_model.eval()
    student_model.train()

    losses, train_time, data_time = [AverageMeter() for _ in range(3)]
    ffn_losses, attn_losses = [AverageMeter() for _ in range(2)]
    st_time = time.time()

    def _update_losses(all_losses, loss, data_size):
        #logging.info("data_size={}".format(data_size))
        if args.distributed:
            loss = reduce_tensor(loss.detach(), args.world_size)
        all_losses.update(loss.item(), data_size)

    num_pca_cka_batch = 0

    student_config = select_config(args.student_model, args.lowercase)
    student_config.print_configs()
    select_times = []
    for i in range(args.type_blocks):
        inside_arr = []
        for j in range(args.num_layers):
            inside_arr.append(0)
        select_times.append(inside_arr)

    for batch_idx, data in enumerate(data_loader):
        data_time.update(time.time() - st_time)
        if use_gpu:
            data = [data_.cuda() for data_ in data]
        token_ids, segment_ids, position_ids, attn_mask, lm_labels = data
        #logging.info(
        #    'teacher_model:{}'.format(
        #        teacher_model))
        with torch.no_grad():
            teacher_outputs = teacher_model(token_ids, segment_ids, position_ids, attn_mask)
        #if args.student_model in nas_bert_models:
        #    student_outputs = student_model(token_ids, segment_ids, position_ids, attn_mask, *args.ffn_arch, is_calc_pca_cka = batch_idx % 100 == 0)
        #else:
        #    student_outputs = student_model(token_ids, segment_ids, position_ids, attn_mask, is_calc_pca_cka = batch_idx % 100 == 0)

        #logging.info('select_time={}'.format(select_times))
        for i in range(args.range2-args.range1+1):
            select_arch = []
            for _ in range(args.student_config.num_layers):
                select_arch.append(i)
            #logging.info("i={}, params={}".format(i, student_model.module.params(select_arch)))
        #logging.info("student_model.module.encoder[0][0].ffn.hidden_size_increment={}".format(student_model.module.encoder[0][0].ffn.hidden_size_increment))
        #logging.info("student_model.module.encoder[0][0].ffn.max_ffn_hidden_size={}".format(
        #    student_model.module.encoder[0][0].ffn.max_ffn_hidden_size))
        select_arch = eval(args.ffn_expr)
        import random
        for i in range(args.student_config.num_layers):
            if eval(args.ffn_expr) == []:
                select_arch.append(random.randint(args.range1, args.range2))
            if args.only2:
                if select_arch[i] < args.type_each_block:
                    select_arch[i] = 0
                else:
                    select_arch[i] = args.type_each_block
            #if args.local_rank == 0:
            #    logging.info("select_times={}, select_arch={}".format(select_times, select_arch))
            if batch_idx % args.calc_freq == 0:
                select_times[select_arch[i]][i] += 1

        head_outputs = []

        def head_hook(module, input, output):
            if batch_idx == 0:
                head_outputs.append(output)

        # Initialize hooks
        i = 0
        for layer in student_model.modules():
            if isinstance(layer, MobileBertAttention) or isinstance(layer, MySupernetBertAttention):
            #if isinstance(layer, MobileBertAttention):
                if args.local_rank == 0: #sum: 144
                    #print("i={}, layer={}".format(i, layer))
                    i += 1
                if hasattr(layer, 'query'):
                    layer.query.register_forward_hook(head_hook)
                if hasattr(layer, 'key'):
                    layer.key.register_forward_hook(head_hook)
                if hasattr(layer, 'value'):
                    layer.value.register_forward_hook(head_hook)
                if hasattr(layer, 'dense'):
                    layer.dense.register_forward_hook(head_hook)

        if get_cka_pair:
            if batch_idx < 1:
                with torch.no_grad():
                    layer_sim, cka_pairs = student_model(token_ids, segment_ids, position_ids, attn_mask, args.type_blocks,
                                                    select_arch,
                                                    is_calc_pca_cka=batch_idx % args.calc_freq == 0, min_pca=args.min_pca,
                                                    get_cka_pair=get_cka_pair)
                    if batch_idx > 0:
                        for i in range(len(old_layer_sim)):
                            old_layer_sim[i] += layer_sim[i]
                    else:
                        old_layer_sim = layer_sim
                    logging.info("layer_sim={}, old_layer_sim={}".format(layer_sim, old_layer_sim))
                if batch_idx < 0:
                    student_outputs = student_model(token_ids, segment_ids, position_ids, attn_mask, args.type_blocks,
                                                    select_arch,
                                                    is_calc_pca_cka=batch_idx % args.calc_freq == 0,
                                                    min_pca=args.min_pca, get_cka_pair=False)
                else:
                    for i in range(len(old_layer_sim)):
                        old_layer_sim[i] /= 1
                    return old_layer_sim, cka_pairs
                #logging.info("student_outputs={}".format(student_outputs))
        else:
            student_outputs = student_model(token_ids, segment_ids, position_ids, attn_mask, args.type_blocks, select_arch,
                                        is_calc_pca_cka=batch_idx % args.calc_freq == 0, min_pca = args.min_pca, get_cka_pair = get_cka_pair)
        loss, attn_loss, ffn_loss, avg_sim = calc_distil_losses(teacher_outputs, student_outputs, use_gpu, is_calc_pca_cka=batch_idx % args.calc_freq == 0)
        #logging.info("avg_sim = {}", avg_sim)
        #if args.local_rank == 0:
        #    logging.info("len(head_outputs)={}".format(len(head_outputs)))
        #    logging.info("attention_condfidence={}".format(attention_condfidence(head_outputs)))
        #    logging.info("attention_condfidence_normalized={}".format(attention_condfidence_normalized(head_outputs)))


        if batch_idx % args.calc_freq == 0:
            sim_another_batch = []
            for _ in range(args.type_blocks):
                sim_another_batch.append([0] * len(select_arch))
            layer_idx = 0
            for arch_id in select_arch:
                sim_another_batch[arch_id][layer_idx] += avg_sim[layer_idx]
                layer_idx += 1

        if batch_idx == 0:
            _, _, _, _, cka_one_batch, pca_one_batch = student_outputs
            sim_one_batch = []
            for _ in range(args.type_blocks):
                sim_one_batch.append([0] * len(select_arch))
            for i in range(args.type_blocks):
                sim_one_batch[i] = np.add(sim_one_batch[i], sim_another_batch[i])
            num_pca_cka_batch += 1
            if args.local_rank == 0:
                logging.info(
                    'batch_idx = {}, num_pca_cka_batch={}, cka_one_batch = {}, pca_one_batch = {}, sim_one_batch = {}, select_times = {}'.format(batch_idx, num_pca_cka_batch, cka_one_batch, pca_one_batch, sim_one_batch, select_times))
        elif batch_idx % args.calc_freq == 0:
        #elif batch_idx < num_pca_cka_batch:
            _, _, _, _, cka_another_batch, pca_another_batch = student_outputs
            for i in range(args.type_blocks):
                cka_one_batch[i] = np.add(cka_one_batch[i], cka_another_batch[i]) # can not use +=
                pca_one_batch[i] = np.add(pca_one_batch[i], pca_another_batch[i])
                sim_one_batch[i] = np.add(sim_one_batch[i], sim_another_batch[i])
            num_pca_cka_batch += 1
            if args.local_rank == 0:
                logging.info(
                    'batch_idx = {}, num_pca_cka_batch={}, cka_another_batch={}, pca_another_batch={}, sim_another_batch = {}, cka_one_batch = {}, pca_one_batch = {}, sim_one_batch = {}, select_times = {}'.format(batch_idx,
                                                                                                          num_pca_cka_batch,
                                                                                                        cka_another_batch,
                                                                                                        pca_another_batch,
                                                                                                        sim_another_batch,
                                                                                                          cka_one_batch,
                                                                                                          pca_one_batch,
                                                                                                        sim_one_batch,
                                                                                                        select_times))

        optimizer.zero_grad()
        if args.fp16:
            with amp.scale_loss(loss, optimizer) as scaled_loss:
                scaled_loss.backward()
            torch.nn.utils.clip_grad_norm_(amp.master_params(optimizer), args.max_grad_norm)
        else:
            loss.backward()
            torch.nn.utils.clip_grad_norm_(student_model.parameters(), args.max_grad_norm)
        optimizer.step()
        scheduler.step()

        #logging.info("token_ids.size={}",token_ids.size)
        _update_losses(losses, loss, token_ids.size(0))
        _update_losses(attn_losses, attn_loss, token_ids.size(0))
        _update_losses(ffn_losses, ffn_loss, token_ids.size(0))

        if use_gpu:
            torch.cuda.synchronize()
        train_time.update(time.time() - st_time)

        if args.local_rank == 0 and \
                (batch_idx == 0 or (batch_idx + 1) % args.disp_freq == 0 or batch_idx + 1 == len(data_loader)):
            lr = scheduler.get_lr()[0]
            logging.info('Epoch: [{}/{}][{}/{}]\t'
                         'LR: {:.2e}\t'
                         'Loss: {loss.val:.4f} ({loss.avg:.4f})\t'
                         'Attn and ffn loss: {attn_loss.val:.4f} {ffn_loss.val:.4f} ({attn_loss.avg:.4f} {ffn_loss.avg:.4f})\t'
                         'Train time: {train_time.val:.4f}s ({train_time.avg:.4f}s)\t'
                         'Load data time: {data_time.val:.4f}s ({data_time.avg:.4f}s)'
                         .format(epoch, args.total_epochs, batch_idx + 1, len(data_loader), lr,
                                 loss=losses, attn_loss=attn_losses, ffn_loss=ffn_losses,
                                 train_time=train_time, data_time=data_time))

        st_time = time.time()

        if args.test == 1:
            break

    for e in range(len(cka_one_batch)):
        for f in range(len(cka_one_batch[e])):
            if select_times[e][f] != 0:
                cka_one_batch[e][f] /= select_times[e][f]
                pca_one_batch[e][f] /= select_times[e][f]
                sim_one_batch[e][f] /= select_times[e][f]

    arr = []
    for i in range(len(pca_one_batch)):
        for j in range(len(pca_one_batch[i])):
            arr.append([pca_one_batch[i][j], i, j])
    arr.sort(key=lambda x: x[0], reverse=True)
    if args.local_rank == 0:
        logging.info("arr={}".format(arr))

    num_add_thistime = args.num_layers * args.average_add_times * args.type_blocks // args.total_epochs
    if args.dynamic == 0:
        num_add_thistime = 0
    if args.only2:
        num_add_thistime /= args.type_each_block
    logging.info("num_add_thistime={}".format(num_add_thistime))
    if num_add_thistime != 0 or args.test == 1:
        total_params = 0
        for _ in range(args.sample_steps):
            sample_arch = []
            for i in range(args.student_config.num_layers):
                sample_arch.append(random.randint(args.range1, args.range2))
                if args.only2:
                    if sample_arch[i] < args.type_each_block:
                        sample_arch[i] = 0
                    else:
                        sample_arch[i] = args.type_each_block
            total_params += student_model.module.params(sample_arch)
        if total_params / args.sample_steps > args.upper_params:
            num_add_thistime = 0
        if args.local_rank == 0:
            logging.info("total_params / args.sample_steps={}".format(total_params / args.sample_steps))
    i = 0
    j = 0
    #logging.info("student_model.params={}".format(student_model.module.params(select_arch)))
    while i < num_add_thistime:
        if student_model.module.add_indices(arr[j][1], arr[j][2]):
            i += 1
        j += 1

    return cka_one_batch, pca_one_batch, sim_one_batch



def calc_distil_losses(teacher_outputs, student_outputs, use_gpu, is_calc_pca_cka):
    teacher_pred_logits, teacher_attn_scores, teacher_attn_outputs, teacher_ffn_outputs = teacher_outputs
    student_pred_logits, student_attn_scores, student_attn_outputs, student_ffn_outputs, _, _ = student_outputs

    def _replace_attn_mask(attn_output):
        replace_values = torch.zeros_like(attn_output)
        if use_gpu:
            replace_values = replace_values.cuda()
        attn_output = torch.where(attn_output <= -1e2, replace_values, attn_output)
        return attn_output
#
    attn_loss, ffn_loss = 0, 0
    if args.new_loss == 0:
        mse_loss = nn.MSELoss()
        ffn_loss += mse_loss(teacher_ffn_outputs[0], student_ffn_outputs[0])
    else:
        pkd = PKD()
        ffn_loss += pkd(teacher_ffn_outputs[0], student_ffn_outputs[0])
    cos_sim = []
    for layer_id in range(args.num_student_layers):
        teacher_layer_id = (layer_id + 1) * args.teacher_interval - 1
        #logging.info('layer_id:{},teacher_layer_id:{},teacher_interval:{}'.format(layer_id, teacher_layer_id, args.teacher_interval))
        #logging.info('teacher_attn_outputs[teacher_layer_id].shape:{},student_attn_outputs[layer_id].shape:{}'.format(
        #    teacher_attn_outputs[teacher_layer_id].shape, student_attn_outputs[layer_id].shape))
        #logging.info('teacher_attn_outputs[teacher_layer_id]:{},student_attn_outputs[layer_id]:{}'.format(
        #    teacher_attn_outputs[teacher_layer_id], student_attn_outputs[layer_id]))
        #logging.info('teacher_ffn_outputs[teacher_layer_id+1].shape:{},student_ffn_outputs[layer_id+1].shape:{}'.format(
        #    teacher_ffn_outputs[teacher_layer_id+1].shape, student_ffn_outputs[layer_id+1].shape))
        #logging.info('teacher_ffn_outputs[teacher_layer_id]:{},student_ffn_outputs[layer_id]:{}'.format(
        #    teacher_ffn_outputs[teacher_layer_id], student_ffn_outputs[layer_id]))
        if args.new_loss == 0:
            attn_loss += mse_loss(teacher_attn_outputs[teacher_layer_id], student_attn_outputs[layer_id])
            ffn_loss += mse_loss(teacher_ffn_outputs[teacher_layer_id + 1], student_ffn_outputs[layer_id + 1])
        else:
            attn_loss += pkd(teacher_attn_outputs[teacher_layer_id], student_attn_outputs[layer_id])
            ffn_loss += pkd(teacher_ffn_outputs[teacher_layer_id + 1], student_ffn_outputs[layer_id + 1])
        if is_calc_pca_cka:
            cos_sim.append(torch.mean(F.cosine_similarity(teacher_ffn_outputs[teacher_layer_id + 1], student_ffn_outputs[layer_id + 1], dim=2)).detach().cpu().numpy())
        #print(torch.mean(F.cosine_similarity(teacher_ffn_outputs[teacher_layer_id + 1], student_ffn_outputs[layer_id + 1], dim=2)).detach().cpu().numpy().shape)
    hidden_loss = attn_loss + ffn_loss
    #logging.info('cos_sim:{}'.format(cos_sim))
    return hidden_loss, attn_loss, ffn_loss, cos_sim


if __name__ == '__main__':
    main()
